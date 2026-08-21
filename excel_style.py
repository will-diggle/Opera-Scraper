"""
One place for how our Excel files look, so every export matches the app:
cream paper, ink headers, Iowan Old Style, and Yes/No columns that go green.
"""

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SERIF = "Iowan Old Style"          # falls back to Excel's default elsewhere
INK = "3B3833"
PAPER = "FAF9F5"
LINE = "E6E3DB"
GREEN = "CFE3D4"
GREEN_ROW = "EDF4EE"


def dress(ws, cols, widths, tick_cols=()):
    """Apply the house style to a sheet that already has a header and rows."""
    last = max(ws.max_row, 2)

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    thin = Side(style="thin", color=LINE)
    for cell in ws[1]:
        cell.font = Font(name=SERIF, bold=True, color="FDFDFB", size=11)
        cell.fill = PatternFill("solid", fgColor=INK)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 26

    for row in ws.iter_rows(min_row=2, max_row=last):
        for cell in row:
            cell.font = Font(name=SERIF, size=11)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.border = Border(bottom=thin)
            if not cell.fill or cell.fill.fgColor.rgb in (None, "00000000"):
                cell.fill = PatternFill("solid", fgColor=PAPER)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{last}"

    # Yes/No dropdowns that turn green when set
    for name in tick_cols:
        if name not in cols:
            continue
        letter = get_column_letter(cols.index(name) + 1)
        dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
        dv.error = "Choose Yes or No"
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{last}")
        ws.conditional_formatting.add(
            f"{letter}2:{letter}{last}",
            FormulaRule(formula=[f'EXACT(${letter}2,"Yes")'],
                        fill=PatternFill("solid", fgColor=GREEN),
                        font=Font(name=SERIF, bold=True, size=11),
                        stopIfTrue=False))

    # tint the whole row once the first tick column is Yes
    if tick_cols and tick_cols[0] in cols:
        first = get_column_letter(cols.index(tick_cols[0]) + 1)
        ws.conditional_formatting.add(
            f"A2:{get_column_letter(len(cols))}{last}",
            FormulaRule(formula=[f'EXACT(${first}2,"Yes")'],
                        fill=PatternFill("solid", fgColor=GREEN_ROW),
                        stopIfTrue=False))
    return ws
