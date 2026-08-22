"""
STAGE 7 - find roles that are NOT YET CAST.

When a house publishes a production but has not signed a singer for a role, the
cast list says "N.N." (nomen nominandum), "TBA", "Besetzung folgt" or similar.
That is a job that exists but has not been advertised - the most useful thing
this whole project can find.

We visit each company's season/repertoire pages, open the production pages, read
the cast list, and note every role with a placeholder instead of a name. Roles
are matched against fach.csv to say which voice type is wanted.

Run:  .venv/bin/python stage7_uncast.py                  (curated list)
      .venv/bin/python stage7_uncast.py --file targets.csv
      .venv/bin/python stage7_uncast.py landestheater    (one company)
Out:  uncast_roles.csv / .xlsx
"""

import csv
import json
import os
import re
import sys
import threading
import time
import unicodedata
from concurrent.futures import ThreadPoolExecutor
from datetime import date
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

COMPANIES_FILE = "companies.csv"
FACH_FILE = "fach.csv"
OUT_CSV = "uncast_roles.csv"
OUT_XLSX = "uncast_roles.xlsx"
CACHE = "uncast_cache.jsonl"

TIMEOUT = 20
PAUSE = 0.6
WORKERS = 8
MAX_SEASON_PAGES = 10     # season/repertoire index pages per company
MAX_PRODUCTIONS = 90      # production pages per company

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    ),
    "Accept-Language": "de,en;q=0.9,it;q=0.7,fr;q=0.7",
}

# Links that lead to the season programme / list of productions.
SEASON_WORDS = [
    "spielplan", "produktion", "repertoire", "programm", "season", "whats-on",
    "what-s-on", "programme", "stagione", "cartelera", "saison", "oper",
    "opera", "vorstellung", "auffuehrung", "productions", "calendar",
]

# A cast list entry with no singer attached yet.
PLACEHOLDER = re.compile(
    r"^(?:"
    r"n\.?\s*n\.?"                     # N.N. / NN / N. N.
    r"|t\.?b\.?[acd]\.?"               # TBA / TBC / TBD
    r"|to be (?:announced|confirmed|cast|advised)"
    r"|cast to be announced|casting in progress"
    # German
    r"|besetzung folgt|wird noch bekannt ?gegeben|noch nicht besetzt"
    r"|steht noch nicht fest|noch offen|in vorbereitung|wird nachgereicht"
    r"|n\.n\. \(.*\)|in k[uü]rze|demn[aä]chst|vakant"
    # Italian
    r"|da definire|in via di definizione|da destinarsi|da nominare"
    r"|sar[aà] comunicato|in fase di definizione"
    # French
    r"|[aà] confirmer|distribution en cours|en cours de distribution"
    r"|[aà] d[eé]finir|prochainement"
    # Spanish / Portuguese
    r"|por confirmar|por determinar|por designar|pendiente de confirmar"
    r"|a confirmar|a designar"
    # Dutch
    r"|nog niet bekend|nader te bepalen|wordt bekendgemaakt|volgt nog"
    # Nordic
    r"|meddelas senare|ej fastst[aä]lld|annonseres senere|kommer senere"
    r"|oplyses senere|ilmoitetaan my[oö]hemmin"
    # Central & Eastern Europe
    r"|do ustalenia|obsada w przygotowaniu|bude up[rř]esn[eě]no"
    r"|p[rř]ipravujeme|k[eé]s[oő]bb|hamarosan"
    r")$", re.I)

# A row of dashes or asterisks might mean "not cast", or might just be a
# separator. Too weak to trust alone - only accepted for a role we recognise
# in an opera the page actually names.
WEAK_MARKER = re.compile(r"^(?:[-–—]{1,}|\*{2,}|\.{3,}|\?{2,}|_{2,})$")

# Lines that are production credits, not singing roles.
CREW = re.compile(
    r"^(?:musikalische leitung|dirigent|inszenierung|regie|b[uü]hne|kost[uü]m|"
    r"licht|dramaturgie|choreograf|chor(?:leitung|einstudierung)?|video|"
    r"conductor|director|design|lighting|costume|set|choreograph|"
    r"maestro concertatore|regia|scene|costumi|luci|"
    r"mise en sc[eè]ne|d[eé]cors|lumi[eè]res|"
    r"orchester|orchestra|statisterie|assistenz|"
    # production metadata that sits in the same list as the cast
    r"dauer|spieldauer|auffuehrungsdauer|l[aä]nge|pause|premiere|termine|"
    r"musik|text|libretto|komponist|autor|nach |fassung|sprache|ort|preise|"
    r"einf[uü]hrung|altersempfehlung|urauff[uü]hrung|besetzung|duration|"
    r"language|running time|prezzi|durata|dur[eé]e|idioma)", re.I)


OPERA_TITLES = set()


def strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFKD", s)
                   if not unicodedata.combining(c))


def norm(s):
    """Loose key for matching role names across languages and spellings."""
    s = strip_accents(s.lower())
    s = s.replace("ss", "s").replace("ae", "a").replace("oe", "o")
    s = s.replace("ue", "u")
    return re.sub(r"[^a-z0-9]+", "", s)


def load_fach():
    table = {}
    if not os.path.exists(FACH_FILE):
        return table
    with open(FACH_FILE, newline="", encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            table.setdefault(norm(r["role"]), []).append(r)
    return table


FACH = load_fach()
OPERA_TITLES.update(norm(r["opera"]) for hits in FACH.values() for r in hits)
_lock = threading.Lock()

# Counters so a run can report how much of the repertoire it actually opened,
# rather than leaving you to guess at its coverage.
STATS = {"companies": 0, "reached": 0, "season_pages": 0, "productions": 0,
         "pages_read": 0, "pages_with_finding": 0}


def bump(key, n=1):
    with _lock:
        STATS[key] += n


class OfflineError(Exception):
    """Raised when the internet has dropped, rather than the sites being dead."""


# Same guard as the job scraper: without it a dropped connection produces a
# run that says "Done. 0 uncast roles" and looks like a real answer.
BREAKER_LIMIT = 25
_misses = 0


def note_home(ok):
    global _misses
    with _lock:
        if ok:
            _misses = 0
            return
        _misses += 1
        if _misses >= BREAKER_LIMIT:
            raise OfflineError(
                f"{_misses} company homepages in a row could not be reached - "
                "your connection looks down. Stopping rather than reporting an "
                "empty result. Nothing unreachable was cached; rerun to continue.")


def log(m):
    print(m, flush=True)


def fetch(url, session):
    try:
        r = session.get(url, timeout=TIMEOUT, headers=HEADERS)
        if r.status_code >= 400:
            return None
        if "html" not in r.headers.get("content-type", ""):
            return None
        return BeautifulSoup(r.content, "lxml")
    except requests.RequestException:
        return None


def same_site(url, root):
    a, b = urlparse(url).netloc.lower(), urlparse(root).netloc.lower()
    a = a[4:] if a.startswith("www.") else a
    b = b[4:] if b.startswith("www.") else b
    return a.endswith(b) or b.endswith(a)


def links_matching(soup, base, words, limit):
    found = {}
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if href.startswith(("mailto:", "tel:", "javascript:", "#")):
            continue
        url = urljoin(base, href).split("#")[0]
        if not url.startswith("http") or not same_site(url, base):
            continue
        blob = (url + " " + a.get_text(" ", strip=True)).lower()
        score = sum(1 for w in words if w in blob)
        if score:
            found[url] = max(score, found.get(url, 0))
    return [u for u, _ in sorted(found.items(), key=lambda kv: -kv[1])][:limit]


def production_title(soup, url):
    for sel in ("h1", "title"):
        el = soup.find(sel)
        if el:
            t = el.get_text(" ", strip=True)
            if t:
                return re.split(r"\s+[|–—-]\s+", t)[0][:90]
    return url.rsplit("/", 1)[-1].replace("-", " ").title()


def read_cast(soup, page_text=""):
    """Return [(role, value)] pairs from a production page.

    Cast lists are laid out as alternating lines - role, then who sings it -
    whether the markup is a table, a definition list or plain divs, so we work
    from the text.
    """
    lines = [l.strip() for l in soup.get_text("\n", strip=True).split("\n")]
    lines = [l for l in lines if l and len(l) < 90]
    pairs = []
    for i in range(len(lines) - 1):
        role, value = lines[i], lines[i + 1]
        if not role or CREW.match(role):
            continue
        if len(role) < 3 or role.endswith((":",)) and len(role) < 4:
            continue
        v = value.strip(" .")
        name = role.rstrip(":").strip()
        if PLACEHOLDER.match(v):
            pairs.append((name, value.strip()))
        elif WEAK_MARKER.match(v) and credible(name, page_text):
            pairs.append((name, value.strip()))
    return pairs


def read_blank_cast(soup, page_text=""):
    """Roles whose singer cell is simply EMPTY.

    Plenty of houses leave the name blank rather than writing N.N., and an
    empty cell vanishes when a page is read as plain text. So this works on
    the markup instead - table rows, definition lists and label/value pairs -
    and reports a role whose value side has nothing in it.

    Empty cells are common in all sorts of tables, so to stay honest we only
    accept a blank when the label is a role we actually recognise.
    """
    found = []

    def consider(label, value):
        label = (label or "").strip().rstrip(":").strip()
        if not label or value.strip():
            return
        if CREW.match(label) or len(label) < 3:
            return
        if credible(label, page_text):
            found.append((label, "(left blank)"))

    for tr in soup.find_all("tr"):
        cells = tr.find_all(["td", "th"], recursive=False) or tr.find_all(["td", "th"])
        if len(cells) == 2:
            consider(cells[0].get_text(" ", strip=True),
                     cells[1].get_text(" ", strip=True))

    for dl in soup.find_all("dl"):
        items = dl.find_all(["dt", "dd"])
        for a, b in zip(items, items[1:]):
            if a.name == "dt" and b.name == "dd":
                consider(a.get_text(" ", strip=True), b.get_text(" ", strip=True))

    # label/value pairs built from divs or spans
    for el in soup.find_all(["li", "div", "p"]):
        kids = [k for k in el.find_all(["span", "div", "strong", "b"],
                                       recursive=False)]
        if len(kids) == 2:
            consider(kids[0].get_text(" ", strip=True),
                     kids[1].get_text(" ", strip=True))
    return found


def credible(role, page_text):
    """Is this really a role on this page, rather than a title in a list?

    Two traps: an opera's name is often also a character's name (Don Giovanni,
    Falstaff, Tosca), and season listings put those titles in tables full of
    empty cells. So a weak signal is only believed when the label is a role we
    know AND the page names the opera that role belongs to.
    """
    key = norm(role)
    if key in OPERA_TITLES:            # "Don Giovanni" as a season entry
        return False
    hits = FACH.get(key, [])
    if not hits:
        return False
    low = strip_accents(page_text.lower())
    return any(strip_accents(h["opera"].lower()) in low for h in hits)


def match_fach(role, page_text):
    """Look the role up in the fach table; prefer the opera named on the page."""
    hits = FACH.get(norm(role), [])
    if not hits:
        return {}
    if len(hits) > 1:
        low = strip_accents(page_text.lower())
        for h in hits:
            if strip_accents(h["opera"].lower()) in low:
                return h
    return hits[0]


# Staff directories and orchestra rosters use the same layout as a cast list,
# so a blank next to a name looks identical to an uncast role. Skip those pages.
NOT_A_CAST_PAGE = re.compile(
    r"equipe|\bteam\b|mitarbeiter|personal|kollegium|orchester|orchestra|"
    r"ensemble-und-team|wer-wir-sind|about-us|ueber-uns|kontakt|impressum|"
    r"verwaltung|leitung|stiftung|freunde|foerderverein|"
    # access and listing pages name several operas at once, so every title
    # on them looks like an uncast role
    r"audio[- ]described|access|barrierefrei|relaxed[- ]performance|surtitl|"
    r"untertitel|gebaerdensprache|touch[- ]tour|schools|gruppen|newsletter|"
    r"abonnement|\babo\b|gutschein|geschenk|preise|tickets?$", re.I)

# A ballet borrows its characters' names from the opera it is based on, so
# "Manon Lescaut" in Die Kameliendame is a dancer, not a singing role.
IS_BALLET = re.compile(
    r"\bballett?\b|ballet|balletto|tanz|\bdance\b|choreograph|"
    r"kameliendame|schwanensee|nussknacker|swan lake|nutcracker|giselle",
    re.I)

# Instrument names appear in orchestra vacancy lists, which are not for singers.
INSTRUMENT = re.compile(
    r"fl[uû]te|floete|hautbois|oboe|clarinette|klarinett|basson|fagott|"
    r"cor anglais|horn|trompet|tromb|tuba|violon|violine|viola|alto solo|"
    r"violoncell|cello|contrebasse|kontrabass|harfe|harpe|percussion|"
    r"schlagzeug|timbales|pauke|klavier|piano|orgel|orgue", re.I)


# Fragments of dates, headings and credits that survive the line pairing.
JUNK_ROLE = re.compile(
    r"^\(|^//|^\d|\)\s*,?$|^[A-ZÄÖÜ ]{4,}$|magazin|koordinator|leitung|"
    r"keine bezeichnung|schauspielerin|schauspieler\b", re.I)


def looks_like_person(text):
    """'Jean-Charles Masurier' is a name; 'Don Ottavio' is a role we may know."""
    if norm(text) in FACH:
        return False
    words = text.split()
    if len(words) != 2:
        return False
    return all(w[:1].isupper() and w[1:].islower() and len(w) > 2 for w in words)


def scan_company(row):
    company = row.get("name", "")
    country = row.get("country", "")
    site = row.get("website", "")
    session = requests.Session()
    out = []

    bump("companies")
    home = fetch(site, session)
    if home is None:
        note_home(False)
        return company, None          # None = never checked, do not cache
    note_home(True)
    bump("reached")

    seasons = links_matching(home, site, SEASON_WORDS, MAX_SEASON_PAGES)
    prod_urls, seen = [], set()
    for s_url in seasons:
        time.sleep(PAUSE)
        s = fetch(s_url, session)
        if s is None:
            continue
        for u in links_matching(s, s_url, SEASON_WORDS, MAX_PRODUCTIONS):
            if u not in seen:
                seen.add(u)
                prod_urls.append(u)
    prod_urls = prod_urls[:MAX_PRODUCTIONS]
    bump("season_pages", len(seasons))
    bump("productions", len(prod_urls))

    for p_url in prod_urls:
        time.sleep(PAUSE)
        p = fetch(p_url, session)
        if p is None:
            continue
        text = p.get_text(" ", strip=True)
        # No pre-filter here. An earlier version skipped any page whose text
        # lacked one of a handful of markers, which (a) threw away pages where
        # the singer cell is simply blank and (b) silently ignored most of the
        # placeholder phrases below. Both passes now run on every page.
        bump("pages_read")
        title = production_title(p, p_url)
        if NOT_A_CAST_PAGE.search(p_url) or NOT_A_CAST_PAGE.search(title):
            continue
        if IS_BALLET.search(title) or IS_BALLET.search(p_url):
            continue
        cast = read_cast(p, text) + read_blank_cast(p, text)
        seen_here = set()
        for role, value in cast:
            if norm(role) in seen_here:
                continue
            seen_here.add(norm(role))
            if INSTRUMENT.search(role) or looks_like_person(role):
                continue
            if JUNK_ROLE.search(role) and norm(role) not in FACH:
                continue
            # a bare dash is too weak on its own - only trust it for a known role
            if value.strip(" .") in {"-", "--", "—", "–"} and norm(role) not in FACH:
                continue
            f = match_fach(role, text)
            out.append({
                "Company": company,
                "Country": country,
                "Production": title,
                "Role": role,
                "Marked as": value,
                # Kloiber's classification, then the looser international
                # reading where it differs. N.A. when the role is not in the
                # table at all - usually a small part or a modern opera.
                "Voice type": f.get("fach", "") or "N.A.",
                "Also called": f.get("fach_common", ""),
                "Voice": f.get("voice", "") or "N.A.",
                "Opera (matched)": f.get("opera", "") or "N.A.",
                "Composer": f.get("composer", "") or "N.A.",
                "Link": p_url,
                "Found on": date.today().isoformat(),
            })

    if out:
        bump("pages_with_finding")

    # a page can repeat a cast block; keep one row per role per production
    uniq, seen_k = [], set()
    for r in out:
        k = (r["Production"], norm(r["Role"]))
        if k not in seen_k:
            seen_k.add(k)
            uniq.append(r)
    if uniq:
        log(f"  ★ {company}: {len(uniq)} uncast role(s)")
    return company, uniq


COLS = ["Company", "Country", "Production", "Role", "Voice type",
        "Also called", "Voice", "Opera (matched)", "Composer", "Marked as",
        "Link", "Found on"]


def write_out(rows):
    rows.sort(key=lambda r: (r["Voice type"] == "N.A.", r["Country"],
                             r["Company"]))
    with open(OUT_CSV, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=COLS, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Uncast roles"
    ws.append(COLS)
    for r in rows:
        ws.append([r.get(c, "") for c in COLS])
    for i, wdt in enumerate([28, 9, 34, 26, 30, 30, 14, 24, 16, 12, 46, 12],
                            start=1):
        ws.column_dimensions[get_column_letter(i)].width = wdt
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{ws.max_row}"
    wb.save(OUT_XLSX)


def main():
    args = sys.argv[1:]
    infile = COMPANIES_FILE
    if "--file" in args:
        i = args.index("--file")
        infile = args[i + 1]
        del args[i:i + 2]

    with open(infile, newline="", encoding="utf-8") as fh:
        companies = [r for r in csv.DictReader(fh) if r.get("website")]
    if args:
        companies = [c for c in companies if args[0].lower() in c["name"].lower()]

    done, rows = set(), []
    if os.path.exists(CACHE):
        with open(CACHE, encoding="utf-8") as fh:
            for line in fh:
                try:
                    d = json.loads(line)
                except json.JSONDecodeError:
                    continue
                done.add(d["company"])
                rows.extend(d["rows"])
    todo = [c for c in companies if c["name"] not in done]

    log(f"{len(FACH)} roles in the fach table")
    log(f"{len(done)} companies already scanned, {len(todo)} to go")

    stopped = ""
    try:
      with open(CACHE, "a", encoding="utf-8") as cache:
        with ThreadPoolExecutor(max_workers=WORKERS) as pool:
            for n, (company, found) in enumerate(
                    pool.map(scan_company, todo), start=1):
                if found is None:                 # unreachable: try again later
                    continue
                cache.write(json.dumps({"company": company, "rows": found},
                                       ensure_ascii=False) + "\n")
                rows.extend(found)
                if n % 25 == 0:
                    cache.flush()
                    log(f"  --- {n}/{len(todo)} companies, {len(rows)} roles ---")
    except OfflineError as e:
        stopped = str(e)
        log(f"\n!! STOPPED: {stopped}")

    write_out(rows)
    withfach = sum(1 for r in rows if r["Voice type"] != "N.A.")
    log("")
    log(f"  companies in the list       {STATS['companies']}")
    log(f"  websites reached            {STATS['reached']}")
    log(f"  season/programme pages read {STATS['season_pages']}")
    log(f"  production pages opened     {STATS['productions']}")
    log(f"  production pages read       {STATS['pages_read']}")
    log(f"  pages with an uncast role   {STATS['pages_with_finding']}")
    log(f"\n{'PARTIAL' if stopped else 'Done'}. {len(rows)} uncast roles "
        f"({withfach} matched to a voice type).")
    log(f"Wrote {OUT_CSV} and {OUT_XLSX}")


if __name__ == "__main__":
    main()
