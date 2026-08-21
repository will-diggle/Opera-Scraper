"""
Opera job scraper.

Reads companies.csv, visits each company's website, looks for pages about
auditions / jobs / young artist programmes, and pulls out anything that looks
like an opportunity. Writes opera_jobs.xlsx (filterable) and opera_jobs.csv.

Run:  .venv/bin/python scraper.py
"""

import csv
import json
import os
import re
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor
from datetime import date
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# --- settings you may want to tweak -----------------------------------------

COMPANIES_FILE = "companies.csv"
OUT_XLSX = "opera_jobs.xlsx"
OUT_CSV = "opera_jobs.csv"

TIMEOUT = 20            # seconds to wait for a page
PAUSE = 1.0             # polite pause between requests to the same site
WORKERS = 24            # how many companies at once (all different sites)
MAX_PAGES_PER_SITE = 8  # how many candidate pages to open per company
MAX_ROWS_PER_COMPANY = 40  # no single site should flood the results

# Words in a LINK that suggest it leads to a jobs/auditions page.
SECTION_WORDS = [
    "audition", "vorsingen", "vorspiel", "audizion", "audicion", "audition",
    "job", "jobs", "career", "careers", "vacanc", "vacature", "employment",
    "opportunit", "recruit", "stellenangebot", "stellen", "offres-emploi",
    "emploi", "lavora-con-noi", "trabaja", "work-with-us", "work-for-us",
    "join-us", "young-artist", "youngartist", "opera-studio", "opernstudio",
    "studio", "chorus", "chor", "ensemble", "apprentice", "fellowship",
    "casting", "engage", "about-us/jobs", "getinvolved",
]

# Words in the TEXT of an item that suggest it is a singer-relevant opening.
SINGER_WORDS = [
    "audition", "singer", "soprano", "mezzo", "contralto", "alto", "tenor",
    "baritone", "bass", "countertenor", "chorus", "choral", "ensemble",
    "young artist", "opera studio", "artist programme", "artist program",
    "apprentice", "fellowship", "cover", "understudy", "soloist",
    "vorsingen", "solist", "chorsänger", "opernchor", "sänger",
    "audizione", "coro", "cantante", "audición", "audition-s",
]

# Words that mean an item is an actual posting rather than a nav link.
POSTING_WORDS = [
    "audition", "apply", "application", "vacancy", "vacancies", "recruiting",
    "now open", "call for", "seeking", "wanted", "opportunity", "position",
    "post of", "deadline", "full-time", "part-time", "fixed-term", "contract",
    "vorsingen", "bewerbung", "stellenausschreibung", "audizione", "audición",
    "concorso", "convocatoria", "engagement",
]

# Generic navigation titles that are never a job posting.
NAV_STOPLIST = {
    "search", "menu", "home", "about", "about us", "contact", "contact us",
    "news", "donate", "support us", "read more", "find out more", "learn more",
    "more info", "back", "next", "previous", "take part", "get involved",
    "sign up", "newsletter", "privacy policy", "cookies", "terms",
    "what's on", "whats on", "tickets", "book now", "our impact", "schools",
    "partnerships", "learning & engagement", "education", "shop", "press",
}

# Words that mean "definitely not a singing job" - used to tag, not to drop.
NON_SINGER_WORDS = [
    "accountant", "marketing", "fundraising", "development officer",
    "electrician", "carpenter", "wardrobe", "usher", "ticketing", "box office",
    "it support", "hr ", "finance", "cleaner", "security", "driver",
    "stage technician", "lighting technician", "sound engineer",
]

DEADLINE_RE = re.compile(
    r"(?:deadline|closing date|closes|apply by|applications close|bewerbungsfrist)"
    r"[:\s]*([0-9]{1,2}\s*[a-zA-Z]{3,9}\s*[0-9]{2,4}|[0-9]{1,2}[./-][0-9]{1,2}[./-][0-9]{2,4})",
    re.I,
)
DATE_RE = re.compile(
    r"\b([0-9]{1,2}\s+(?:January|February|March|April|May|June|July|August|"
    r"September|October|November|December)\s+[0-9]{4}"
    r"|[0-9]{1,2}[./-][0-9]{1,2}[./-][0-9]{2,4}"
    r"|[0-9]{4}-[0-9]{2}-[0-9]{2})\b",
    re.I,
)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    ),
    "Accept-Language": "en,de;q=0.8,fr;q=0.6,it;q=0.6",
}


def log(msg):
    print(msg, flush=True)


class OfflineError(Exception):
    """Raised when we conclude the internet has dropped, not the website."""


# If this many companies fail to connect one after another, it is almost
# certainly our own connection that has gone, not thousands of dead websites.
# Without this guard a dropped wifi produces a run that says "Done" and quietly
# reports every remaining company as unreachable.
BREAKER_LIMIT = 40
_consecutive_conn_fail = 0
_breaker_lock = threading.Lock()


def note_connection(failed):
    global _consecutive_conn_fail
    with _breaker_lock:
        if failed:
            _consecutive_conn_fail += 1
            if _consecutive_conn_fail >= BREAKER_LIMIT:
                raise OfflineError(
                    f"{_consecutive_conn_fail} sites in a row could not be "
                    "reached - your internet connection looks down. Stopping so "
                    "the results are not silently wrong. Everything checked so "
                    "far is saved; rerun the same command to carry on.")
        else:
            _consecutive_conn_fail = 0


def fetch(url, session, retry=True):
    """Get a page. Returns (soup, final_url) or (None, reason)."""
    try:
        r = session.get(url, timeout=TIMEOUT, headers=HEADERS, allow_redirects=True)
        note_connection(False)
        if r.status_code >= 400:
            return None, f"HTTP {r.status_code}"
        ctype = r.headers.get("content-type", "")
        if "html" not in ctype and "xml" not in ctype:
            return None, f"not html ({ctype.split(';')[0]})"
        # Pass raw bytes so lxml reads the page's own charset - otherwise
        # accented and German/Italian pages come out as mojibake.
        return BeautifulSoup(r.content, "lxml"), r.url
    except (requests.ConnectionError, requests.Timeout) as e:
        if retry:
            time.sleep(3)
            return fetch(url, session, retry=False)
        note_connection(True)
        return None, type(e).__name__
    except requests.RequestException as e:
        note_connection(False)
        return None, type(e).__name__


def same_site(url, root):
    a, b = urlparse(url).netloc.lower(), urlparse(root).netloc.lower()
    a = a[4:] if a.startswith("www.") else a
    b = b[4:] if b.startswith("www.") else b
    return a.endswith(b) or b.endswith(a)


def find_section_links(soup, base_url):
    """Links on a page that look like they lead to jobs/auditions."""
    hits = {}
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if href.startswith(("mailto:", "tel:", "javascript:", "#")):
            continue
        url = urljoin(base_url, href)
        if not url.startswith("http") or not same_site(url, base_url):
            continue
        haystack = (url + " " + a.get_text(" ", strip=True)).lower()
        score = sum(1 for w in SECTION_WORDS if w in haystack)
        if score:
            hits[url.split("#")[0]] = max(score, hits.get(url.split("#")[0], 0))
    return [u for u, _ in sorted(hits.items(), key=lambda kv: -kv[1])]


def _word_re(words):
    """Match whole words only, so 'bass' doesn't fire on 'bassoon'."""
    return re.compile(r"(?<![a-z])(?:" + "|".join(re.escape(w.strip())
                      for w in words) + r")(?![a-z])", re.I)


SINGER_RE = _word_re(SINGER_WORDS)
POSTING_RE = _word_re(POSTING_WORDS)
NON_SINGER_RE = _word_re(NON_SINGER_WORDS)


def classify(text):
    if SINGER_RE.search(text):
        return "Singer"
    if NON_SINGER_RE.search(text):
        return "Other staff"
    return "Unclear"


def looks_like_posting(title, strong_page=True):
    t = title.strip().lower().strip(" -–—:")
    if t in NAV_STOPLIST or len(t) < 8:
        return False
    if t.startswith(("read more", "find out", "learn more", "more about")):
        return False
    if not strong_page:
        # e.g. an "ensemble" page - only keep clear calls for applicants.
        return bool(POSTING_RE.search(title))
    return bool(SINGER_RE.search(title) or POSTING_RE.search(title))


# A page whose own URL says "jobs/auditions" is trusted; anywhere else
# (an ensemble or studio page) we demand stronger wording in the item itself,
# otherwise we scrape cast lists and singer biographies.
STRONG_PAGE_RE = re.compile(
    r"audition|vorsingen|vorspiel|audizion|audicion|job|career|vacan|vacature"
    r"|employ|recruit|stellen|emploi|lavora|trabaj|opportunit|bewerb|concorso"
    r"|work-with-us|work-for-us|join-us", re.I)


def extract_items(soup, page_url, company, country):
    """Pull candidate openings off a jobs/auditions page."""
    page_text = soup.get_text(" ", strip=True)
    strong_page = bool(STRONG_PAGE_RE.search(page_url))
    items = []
    seen = set()

    # 1. Links on the page that read like an individual posting.
    for a in soup.find_all("a", href=True):
        title = a.get_text(" ", strip=True)
        if not (6 <= len(title) <= 160):
            continue
        href = urljoin(page_url, a["href"].strip())
        if not href.startswith("http"):
            continue
        if not looks_like_posting(title, strong_page):
            continue
        key = title.lower()
        if key in seen:
            continue
        seen.add(key)
        items.append((title, href, title))

    # 2. Headings that read like a posting, when links gave us nothing.
    if not items:
        for h in soup.find_all(["h1", "h2", "h3", "h4", "li"]):
            title = h.get_text(" ", strip=True)
            if not (10 <= len(title) <= 160):
                continue
            if not looks_like_posting(title, strong_page):
                continue
            key = title.lower()
            if key in seen:
                continue
            seen.add(key)
            items.append((title, page_url, title))

    deadline = ""
    m = DEADLINE_RE.search(page_text)
    if m:
        deadline = m.group(1).strip()

    rows = []
    for title, link, blob in items:
        d = deadline
        if not d:
            near = DATE_RE.search(blob)
            d = near.group(1) if near else ""
        rows.append({
            "Company": company,
            "Country": country,
            "Role / posting": title,
            "Type": classify(blob),
            "Deadline found": d,
            "Link": link,
            "Found on page": page_url,
            "Date checked": date.today().isoformat(),
            "Social accounts": "",
            "Status": "",
            "Notes": "",
        })
    return rows


def scrape_company(row):
    company, country, site = row["name"], row["country"], row["website"]
    session = requests.Session()
    results, notes = [], ""

    soup, final = fetch(site, session)
    if soup is None:
        log(f"  ✗ {company}: homepage failed ({final})")
        return [], {"Company": company, "Country": country, "Website": site,
                    "Result": f"could not open homepage ({final})"}

    candidates = find_section_links(soup, final)[:MAX_PAGES_PER_SITE]
    if not candidates:
        notes = "no jobs/auditions link found on homepage"

    visited = set()
    for url in candidates:
        if url in visited:
            continue
        visited.add(url)
        time.sleep(PAUSE)
        s, f = fetch(url, session)
        if s is None:
            continue
        results.extend(extract_items(s, f, company, country))

    # de-duplicate by (title, link)
    uniq, seen = [], set()
    for r in results:
        k = (r["Role / posting"].lower(), r["Link"])
        if k not in seen:
            seen.add(k)
            uniq.append(r)

    if len(uniq) > MAX_ROWS_PER_COMPANY:
        log(f"  ! {company}: {len(uniq)} items, keeping first "
            f"{MAX_ROWS_PER_COMPANY} - looks like a general careers site")
        uniq = uniq[:MAX_ROWS_PER_COMPANY]

    # carry any social accounts we know about, so you can click through and
    # check their feeds by hand - social sites block automated reading.
    socials = " | ".join(filter(None, [row.get(k, "") for k in
                         ("facebook", "instagram", "twitter_x", "linkedin")]))
    for r in uniq:
        r["Social accounts"] = socials

    log(f"  ✓ {company}: {len(uniq)} item(s) from {len(visited)} page(s)")
    return uniq, {"Company": company, "Country": country, "Website": site,
                  "Result": notes or f"{len(uniq)} items from {len(visited)} pages",
                  "Socials": socials}


def write_xlsx(rows, coverage):
    wb = Workbook()
    ws = wb.active
    ws.title = "Opportunities"
    cols = ["Company", "Country", "Role / posting", "Type", "Deadline found",
            "Link", "Found on page", "Date checked", "Social accounts",
            "Status", "Notes"]
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c, "") for c in cols])

    widths = [28, 12, 60, 12, 18, 50, 50, 13, 45, 12, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"

    ws2 = wb.create_sheet("Coverage")
    ws2.append(["Company", "Country", "Website", "Result", "Socials"])
    for c in coverage:
        ws2.append([c["Company"], c["Country"], c["Website"], c["Result"],
                    c.get("Socials", "")])
    for i, w in enumerate([28, 12, 45, 45, 45], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for c in ws2[1]:
        c.font = Font(bold=True)
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:E{ws2.max_row}"

    wb.save(OUT_XLSX)


def main():
    args = sys.argv[1:]
    infile, out_prefix = COMPANIES_FILE, "opera_jobs"
    if "--file" in args:
        i = args.index("--file")
        infile = args[i + 1]
        out_prefix = "jobs_" + infile.rsplit(".", 1)[0]
        del args[i:i + 2]

    global OUT_XLSX, OUT_CSV
    OUT_XLSX, OUT_CSV = out_prefix + ".xlsx", out_prefix + ".csv"

    with open(infile, newline="", encoding="utf-8") as fh:
        companies = [r for r in csv.DictReader(fh) if r.get("website")]

    only = args[0].lower() if args else None
    if only:
        companies = [c for c in companies if only in c["name"].lower()]

    # Big runs take hours, so remember what we've already checked. Stop with
    # Ctrl-C and rerun to carry on; delete the .cache file to start fresh.
    cache_path = out_prefix + "_cache.jsonl"
    all_rows, coverage, done = [], [], set()
    if os.path.exists(cache_path):
        with open(cache_path, encoding="utf-8") as fh:
            for line in fh:
                try:
                    d = json.loads(line)
                except json.JSONDecodeError:
                    continue
                done.add(d["company"])
                all_rows.extend(d["rows"])
                coverage.append(d["coverage"])
    todo = [c for c in companies if c["name"] not in done]

    log(f"{len(done)} already checked, {len(todo)} to go "
        f"(of {len(companies)} companies)")
    start, n, stopped = time.time(), 0, ""
    try:
      with open(cache_path, "a", encoding="utf-8") as cache:
        with ThreadPoolExecutor(max_workers=WORKERS) as pool:
            for company, (rows, cov) in zip(
                    [c["name"] for c in todo],
                    pool.map(scrape_company, todo)):
                # A connection failure usually means OUR network hiccupped,
                # so don't record it as checked - let a rerun try again.
                transient = any(w in cov.get("Result", "") for w in
                                ("ConnectionError", "Timeout"))
                if not transient:
                    cache.write(json.dumps(
                        {"company": company, "rows": rows, "coverage": cov},
                        ensure_ascii=False) + "\n")
                all_rows.extend(rows)
                coverage.append(cov)
                n += 1
                if n % 200 == 0:
                    cache.flush()
                    rate = n / (time.time() - start)
                    log(f"  --- {n}/{len(todo)} companies, "
                        f"{len(all_rows)} rows, "
                        f"~{(len(todo) - n) / rate / 60:.0f} min left ---")
    except (OfflineError, KeyboardInterrupt) as e:
        stopped = str(e) or "stopped by you (Ctrl-C)"
        log(f"\n!! STOPPED: {stopped}")

    all_rows.sort(key=lambda r: (r["Type"] != "Singer", r["Country"], r["Company"]))

    with open(OUT_CSV, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=list(all_rows[0].keys()) if all_rows else
                           ["Company", "Country", "Role / posting", "Type",
                            "Deadline found", "Link", "Found on page",
                            "Date checked", "Social accounts", "Status",
                            "Notes"])
        w.writeheader()
        w.writerows(all_rows)

    write_xlsx(all_rows, coverage)
    singers = sum(1 for r in all_rows if r["Type"] == "Singer")
    log(f"\n{'PARTIAL' if stopped else 'Done'}. "
        f"{len(all_rows)} rows ({singers} tagged 'Singer').")
    log(f"Wrote {OUT_XLSX} and {OUT_CSV}")
    if stopped:
        log("This run did NOT finish - rerun the same command to continue.")


if __name__ == "__main__":
    main()
