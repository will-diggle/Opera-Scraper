"""
Your shortlist: the jobs you have actually decided to go for.

Everything else in this project is a firehose the scrapers refill. This file is
yours - what you saved, whether you have applied, whether you have emailed, and
your own notes. Nothing here is ever overwritten by a scrape.

Kept in saved_jobs.json for the app's own use; the file you actually open is
saved_jobs.xlsx, with Yes-No columns that turn green.
"""

import json
import os
from datetime import date

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
STORE = os.path.join(HERE, "saved_jobs.json")
XLSX = os.path.join(HERE, "saved_jobs.xlsx")

FIELDS = ["id", "Company", "Country", "City", "Role", "Voice type",
          "Deadline", "Link", "Saved on", "Applied", "Emailed", "Replied",
          "Notes"]


def row_id(row):
    """Same posting from two scrapes must land on the same shortlist entry."""
    return "|".join([
        (row.get("Company") or "").strip().lower(),
        (row.get("Role / posting") or row.get("Role") or "").strip().lower()[:60],
        (row.get("Link") or "").strip().lower(),
    ])


def load():
    if not os.path.exists(STORE):
        return []
    try:
        with open(STORE, encoding="utf-8") as fh:
            return json.load(fh)
    except (json.JSONDecodeError, OSError):
        return []


def save_all(rows):
    with open(STORE, "w", encoding="utf-8") as fh:
        json.dump([{k: r.get(k, "") for k in FIELDS} for r in rows],
                  fh, ensure_ascii=False, indent=1)


def add(row):
    """Add one posting to the shortlist. Returns True if it was new."""
    rows = load()
    rid = row_id(row)
    if any(r["id"] == rid for r in rows):
        return False
    rows.append({
        "id": rid,
        "Company": row.get("Company", ""),
        "Country": row.get("Country", ""),
        "City": row.get("City", ""),
        "Role": row.get("Role / posting") or row.get("Role", ""),
        "Voice type": row.get("Voice type", ""),
        "Deadline": row.get("Deadline found") or row.get("Deadline", ""),
        "Link": row.get("Link", ""),
        "Saved on": date.today().isoformat(),
        "Applied": "No",
        "Emailed": "No",
        "Replied": "No",
        "Notes": "",
    })
    save_all(rows)
    return True


def remove(rid):
    rows = load()
    kept = [r for r in rows if r["id"] != rid]
    save_all(kept)
    return len(kept) != len(rows)


def update(rid, field, value):
    """Tick a box or edit a note."""
    if field not in ("Applied", "Emailed", "Replied", "Notes"):
        return False
    rows = load()
    for r in rows:
        if r["id"] == rid:
            r[field] = value
            save_all(rows)
            return True
    return False


def export_xlsx():
    """Write the shortlist as a working Excel tracker.

    Applied / Emailed / Replied are Yes-No dropdowns. A cell set to Yes turns
    green, and once you have applied the whole row is tinted, so the sheet
    shows progress at a glance.
    """
    rows = load()
    cols = [c for c in FIELDS if c != "id"]

    wb = Workbook()
    ws = wb.active
    ws.title = "My applications"
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c, "") for c in cols])

    widths = [26, 10, 15, 42, 26, 14, 46, 11, 10, 10, 10, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="3B3833")
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    last = max(ws.max_row, 2)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{last}"

    # Yes / No dropdowns on the three tick columns
    first_tick = cols.index("Applied") + 1
    for offset in range(3):
        letter = get_column_letter(first_tick + offset)
        dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
        dv.error = "Choose Yes or No"
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{last}")
        green = PatternFill("solid", fgColor="C6EFCE")
        ws.conditional_formatting.add(
            f"{letter}2:{letter}{last}",
            FormulaRule(formula=[f'EXACT(${letter}2,"Yes")'], fill=green,
                        stopIfTrue=False))

    # whole row tinted once applied
    applied_col = get_column_letter(first_tick)
    ws.conditional_formatting.add(
        f"A2:{get_column_letter(len(cols))}{last}",
        FormulaRule(formula=[f'EXACT(${applied_col}2,"Yes")'],
                    fill=PatternFill("solid", fgColor="EAF6EC"),
                    stopIfTrue=False))

    wb.save(XLSX)
    return XLSX, len(rows)


if __name__ == "__main__":
    path, n = export_xlsx()
    print(f"Wrote {path} ({n} saved jobs)")
