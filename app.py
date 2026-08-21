"""
The Opera Jobs app.

A small website that runs on your own Mac. Start it, open the page, and you get
a search box, filters and a button that re-runs the scrapers. No spreadsheets
to hunt through.

Start it with:   .venv/bin/python app.py
Then open:       http://localhost:5055
Stop it with:    Ctrl-C in the terminal
"""

import csv
import io
import os
import subprocess
import sys
import threading
from datetime import datetime

from flask import (Flask, jsonify, render_template_string, request,
                   send_file)

import tracker
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from freshness import assess
from stage5_prioritise import categorise

HERE = os.path.dirname(os.path.abspath(__file__))
PY = os.path.join(HERE, ".venv", "bin", "python")

# The jobs files we know about, best first.
JOB_FILES = ["jobs_targets.csv", "opera_jobs.csv"]

app = Flask(__name__)

# --- running a scrape in the background -------------------------------------

state = {"running": False, "job": "", "log": [], "started": ""}
_lock = threading.Lock()

TASKS = {
    "curated": {
        "label": "Check my curated list (222 houses, ~10 min)",
        "cmd": [PY, "scraper.py"],
    },
    "full": {
        "label": "Full Operabase sweep (12,000 companies, ~90 min)",
        "cmd": [PY, "scraper.py", "--file", "targets.csv"],
    },
    "prioritise": {
        "label": "Re-grade results (instant)",
        "cmd": [PY, "stage5_prioritise.py"],
    },
    "refresh_directory": {
        "label": "Refresh company directory from Operabase (~90 min)",
        "cmd": [PY, "stage2_details.py"],
    },
}


def run_task(key):
    cmd = TASKS[key]["cmd"]
    with _lock:
        state.update(running=True, job=TASKS[key]["label"], log=[],
                     started=datetime.now().strftime("%H:%M"))
    try:
        proc = subprocess.Popen(cmd, cwd=HERE, stdout=subprocess.PIPE,
                                stderr=subprocess.STDOUT, text=True,
                                bufsize=1)
        for line in proc.stdout:
            with _lock:
                state["log"].append(line.rstrip())
                del state["log"][:-400]      # keep the last 400 lines
        proc.wait()
        with _lock:
            state["log"].append(
                "FINISHED" if proc.returncode == 0
                else f"STOPPED (exit code {proc.returncode})")
    except Exception as e:                    # noqa: BLE001 - show user anything
        with _lock:
            state["log"].append(f"ERROR: {e}")
    finally:
        with _lock:
            state["running"] = False


# --- reading the results ----------------------------------------------------

def load_jobs():
    """Read whichever results file exists, newest first."""
    rows = []
    for name in JOB_FILES:
        path = os.path.join(HERE, name)
        if not os.path.exists(path):
            continue
        with open(path, newline="", encoding="utf-8") as fh:
            for r in csv.DictReader(fh):
                r["_source"] = name
                rows.append(r)
    # add the Priority grade if stage 5 has been run
    meta = {}
    tpath = os.path.join(HERE, "targets.csv")
    if os.path.exists(tpath):
        with open(tpath, newline="", encoding="utf-8") as fh:
            meta = {r["name"]: r for r in csv.DictReader(fh)}
    for r in rows:
        m = meta.get(r.get("Company", ""), {})
        r["City"] = m.get("city", "")
        r["Posted"], r["Freshness"] = assess(r)
        r["Priority"] = categorise(r.get("Company", ""), m.get("org_type", ""))
    return rows


PAGE = """
<!doctype html><meta charset="utf-8"><title>Opera Jobs</title>
<style>
 :root{--bg:#faf9f5;--panel:#ffffff;--ink:#1f1e1b;--line:#e6e3db;
        --accent:#3b3833;--soft:#736f66;--tint:#f0eee7;
        /* Iowan Old Style throughout - a Venetian oldstyle, warm and
           bookish. Falls back to Palatino, then Georgia, off this Mac. */
        --serif:"Iowan Old Style","Palatino Linotype",Palatino,
                "Book Antiqua",Georgia,serif;
        --sans:var(--serif)}
 *{box-sizing:border-box}
 body{margin:0;font:15.5px/1.55 var(--serif);
      background:var(--bg);color:var(--ink);-webkit-font-smoothing:antialiased}
 header{background:transparent;color:var(--ink);padding:34px 24px 20px;
        max-width:1500px;border-bottom:1px solid var(--line)}
 header .eyebrow{font:600 11px/1 var(--serif);letter-spacing:.13em;
        text-transform:uppercase;color:var(--soft)}
 header h1{margin:10px 0 0;font:400 37px/1.1 var(--serif);
        letter-spacing:-.008em}
 header p{margin:10px 0 0;color:var(--soft);font:16px/1.5 var(--serif);
        max-width:60ch}
 .wrap{padding:20px 24px;max-width:1500px}
 .panel > strong{font:600 15.5px/1.2 var(--serif)}
 td .co, td b{font:600 15px/1.3 var(--serif)}
 .panel{background:var(--panel);border:1px solid var(--line);border-radius:12px;
        padding:18px;margin-bottom:16px}
 button{font:inherit;padding:8px 14px;border-radius:9px;border:1px solid var(--accent);
        background:var(--accent);color:#fdfdfb;cursor:pointer;margin:3px 4px 3px 0}
 button.ghost{background:transparent;color:var(--accent);border-color:var(--line)}
 button.ghost:hover{background:var(--tint)}
 button:disabled{opacity:.45;cursor:not-allowed}
 input,select{font:inherit;padding:8px 10px;border:1px solid var(--line);
              border-radius:9px;background:var(--panel);color:var(--ink)}
 input[type=search]{width:320px}
 table{border-collapse:collapse;width:100%;background:var(--panel);
       font-size:14px;font-variant-numeric:tabular-nums;border:1px solid var(--line);border-radius:12px}
 th,td{text-align:left;padding:8px 10px;border-bottom:1px solid var(--line);
       vertical-align:top}
 th{background:var(--tint);position:sticky;top:0;cursor:pointer;
       white-space:nowrap;font:600 11.5px/1 var(--serif);
       letter-spacing:.07em;text-transform:uppercase;color:var(--soft);
       padding:11px 10px}
 tr:hover td{background:var(--bg)}
 a{color:var(--accent)}
 .tag{font-size:11px;padding:2px 7px;border-radius:99px;background:var(--tint);
      white-space:nowrap}
 .Singer{background:#e4e0d6}.Unclear{background:var(--tint)}
 .log{background:#2a2823;color:#e8e4da;font:12px/1.45 ui-monospace,Menlo,monospace;
      padding:12px;border-radius:10px;height:190px;overflow:auto;white-space:pre-wrap}
 .muted{color:var(--soft);font-size:13px}
 .multi{position:relative;display:inline-block}
 .pop{display:none;position:absolute;z-index:20;top:105%;left:0;background:var(--panel);
      border:1px solid var(--line);border-radius:12px;padding:8px;min-width:230px;
      max-height:280px;overflow:auto;box-shadow:0 6px 20px rgba(0,0,0,.13)}
 .pop.open{display:block}
 .pop label{display:block;padding:3px 4px;font-size:13.5px;cursor:pointer;
            white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
 .pop label:hover{background:#f0ece6}
 .pop input{margin-right:7px}
 .pop .tools{display:flex;gap:6px;margin-bottom:6px;position:sticky;top:0;background:var(--panel)}
 .pop .tools input[type=search]{width:100%;padding:5px}
 td.pick,th.pick{width:34px;text-align:center}
 td.pick input{width:16px;height:16px;cursor:pointer;accent-color:var(--accent)}
 td.tick{text-align:center}
 td.tick input{width:16px;height:16px;accent-color:#1d6b4f}
 tr.done td{background:#eef4ec}
 .notes{width:100%;border:1px solid var(--line);border-radius:8px;padding:6px;
        font:inherit;font-size:13px}
 .fresh-Outofdate{color:#a33}.fresh-Recent{color:#2a7}.fresh-Upcoming{color:#7a5c1e;font-weight:600}
</style>
<header>
  <div class="eyebrow">Local &middot; runs on this Mac</div>
  <h1>Opera Jobs</h1>
  <p>Auditions, chorus vacancies and young artist programmes, gathered from
     company websites. Tick anything worth chasing to build your shortlist.</p>
</header>
<div class="wrap">

<div class="panel">
  <strong>Re-scrape</strong>
  <div style="margin-top:8px">
    {% for key, t in tasks.items() %}
      <button onclick="startJob('{{key}}')" class="{{ '' if key=='curated' else 'ghost' }}">
        {{ t.label }}
      </button>
    {% endfor %}
  </div>
  <div id="status" class="muted" style="margin-top:10px"></div>
  <div class="log" id="log" style="display:none"></div>
</div>

<div class="panel">
  <button id="tab_all" onclick="showView('all')">All findings</button>
  <button id="tab_saved" class="ghost" onclick="showView('saved')">
    My shortlist (<span id="savedcount">0</span>)</button>
  <button class="ghost" style="float:right"
          onclick="location='/api/saved/export'">Export shortlist to Excel</button>
</div>

<div class="panel" id="filters">
  <input type="search" id="q" placeholder="Search role, company, city…"
         oninput="render()">
  <select id="type" onchange="render()">
    <option value="">Any type</option>
    <option value="Singer" selected>Singer only</option>
    <option value="Unclear">Unclear</option>
  </select>
  <select id="fresh" onchange="render()">
    <option value="hideold" selected>Hide out-of-date</option>
    <option value="">Show everything</option>
    <option value="Recent">Recent only</option>
    <option value="Upcoming">Upcoming deadlines</option>
  </select>
  <span class="multi"><button class="ghost" onclick="openPanel('priority')"
        id="b_priority">Priority: all</button><div class="pop" id="p_priority"></div></span>
  <span class="multi"><button class="ghost" onclick="openPanel('country')"
        id="b_country">Country: all</button><div class="pop" id="p_country"></div></span>
  <span class="multi"><button class="ghost" onclick="openPanel('company')"
        id="b_company">Company: all</button><div class="pop" id="p_company"></div></span>
  <label class="muted" style="margin-left:8px">
    <input type="checkbox" id="deadline" onchange="render()"> has a deadline
  </label>
  <button onclick="exportExcel()" style="float:right">Export to Excel</button>
  <button class="ghost" onclick="clearFilters()" style="float:right">Clear filters</button>
  <span id="count" class="muted" style="margin-left:10px"></span>
</div>

<table id="saved_tbl" style="display:none">
  <thead><tr>
    <th>Company</th><th>Role</th><th>Voice type</th><th>Deadline</th>
    <th>Applied</th><th>Emailed</th><th>Replied</th><th>Notes</th>
    <th>Link</th><th></th>
  </tr></thead>
  <tbody></tbody>
</table>

<table id="tbl">
  <thead><tr>
    <th class="pick" title="Tick to add to your shortlist">Save</th>
    <th onclick="sortBy('Priority')">Priority</th>
    <th onclick="sortBy('Company')">Company</th>
    <th onclick="sortBy('Country')">Country</th>
    <th onclick="sortBy('Role / posting')">Role / posting</th>
    <th onclick="sortBy('Type')">Type</th>
    <th onclick="sortBy('Deadline found')">Deadline</th>
    <th onclick="sortBy('Posted')">Posted</th>
    <th onclick="sortBy('Date checked')">Checked</th>
    <th>Links</th>
  </tr></thead>
  <tbody></tbody>
</table>
</div>

<script>
let ROWS = [], SHOWN = [], sortKey = "Company", sortDir = 1;
let SAVEDROWS = [], VIEW = "all";
const SAVED = new Set();

function rid(r){
  return [(r.Company||"").trim().toLowerCase(),
          (r["Role / posting"]||r.Role||"").trim().toLowerCase().slice(0,60),
          (r.Link||"").trim().toLowerCase()].join("|");
}

function loadSaved(){
  return fetch("/api/saved").then(r => r.json()).then(d => {
    SAVEDROWS = d;
    SAVED.clear(); d.forEach(r => SAVED.add(r.id));
    document.getElementById("savedcount").textContent = d.length;
    if (VIEW === "saved") drawSaved(); else render();
  });
}

function toggleSave(json){
  const r = JSON.parse(json), id = rid(r);
  const url = SAVED.has(id) ? "/api/saved/remove" : "/api/saved/add";
  const body = SAVED.has(id) ? {id} : r;
  fetch(url, {method:"POST", headers:{"Content-Type":"application/json"},
              body: JSON.stringify(body)}).then(() => loadSaved());
}

function showView(v){
  VIEW = v;
  document.getElementById("tab_all").className = v==="all" ? "" : "ghost";
  document.getElementById("tab_saved").className = v==="saved" ? "" : "ghost";
  document.getElementById("filters").style.display = v==="all" ? "" : "none";
  document.getElementById("tbl").style.display = v==="all" ? "" : "none";
  document.getElementById("saved_tbl").style.display = v==="saved" ? "" : "none";
  v === "saved" ? drawSaved() : render();
}

function setField(id, field, value){
  fetch("/api/saved/update", {method:"POST",
    headers:{"Content-Type":"application/json"},
    body: JSON.stringify({id, field, value})}).then(() => loadSaved());
}

function drawSaved(){
  const tb = document.querySelector("#saved_tbl tbody");
  if (!SAVEDROWS.length){
    tb.innerHTML = `<tr><td colspan="10" class="muted"
      style="padding:26px">Nothing saved yet. Click the star beside any
      finding to add it here.</td></tr>`;
    return;
  }
  const tick = (r, f) => `<td class="tick"><input type="checkbox"
      ${r[f]==="Yes"?"checked":""}
      onchange="setField('${r.id.replace(/'/g,"\\'")}','${f}',
                         this.checked?'Yes':'No')"></td>`;
  tb.innerHTML = SAVEDROWS.map(r => `
    <tr class="${r.Applied==="Yes"?"done":""}">
      <td><b>${esc(r.Company)}</b>${r.City?`<div class="muted">${esc(r.City)}</div>`:""}</td>
      <td>${esc(r.Role)}</td>
      <td>${esc(r["Voice type"]||"")}</td>
      <td>${esc(r.Deadline||"")}</td>
      ${tick(r,"Applied")}${tick(r,"Emailed")}${tick(r,"Replied")}
      <td><input class="notes" value="${esc(r.Notes||"")}"
           placeholder="date sent, who you wrote to…"
           onchange="setField('${r.id.replace(/'/g,"\\'")}','Notes',this.value)"></td>
      <td><a href="${esc(r.Link)}" target="_blank" rel="noopener">open</a></td>
      <td><button class="ghost" onclick="if(confirm('Remove from shortlist?'))
            fetch('/api/saved/remove',{method:'POST',
            headers:{'Content-Type':'application/json'},
            body:JSON.stringify({id:'${r.id.replace(/'/g,"\\'")}'})})
            .then(loadSaved)">remove</button></td>
    </tr>`).join("");
}

fetch("/api/jobs").then(r => r.json()).then(d => {
  ROWS = d;
  buildMulti("priority", [...new Set(d.map(r => r.Priority))].sort());
  buildMulti("country",  [...new Set(d.map(r => r.Country))].sort());
  buildMulti("company",  [...new Set(d.map(r => r.Company))].sort());
  render();
  loadSaved();
});

const SEL = {priority:new Set(), country:new Set(), company:new Set()};
const OPTS = {};

function buildMulti(name, vals) {
  OPTS[name] = vals.filter(Boolean);
  drawMulti(name, "");
}
function drawMulti(name, filterText) {
  const pop = document.getElementById("p_" + name);
  const shown = OPTS[name].filter(v =>
    v.toLowerCase().includes((filterText||"").toLowerCase()));
  pop.innerHTML =
    `<div class="tools"><input type="search" placeholder="find…"
        oninput="drawMulti('${name}', this.value)" value="${esc(filterText||"")}"></div>` +
    `<label><input type="checkbox" onchange="toggleAll('${name}', this.checked)"
        ${SEL[name].size===0?"checked":""}> <b>All</b></label>` +
    shown.map(v => `<label><input type="checkbox" value="${esc(v)}"
        onchange="pick('${name}', this)" ${SEL[name].has(v)?"checked":""}>
        ${esc(v)}</label>`).join("");
}
function openPanel(name) {
  document.querySelectorAll(".pop").forEach(p =>
    p.classList.toggle("open", p.id === "p_" + name && !p.classList.contains("open")));
  event.stopPropagation();
}
document.addEventListener("click", e => {
  if (!e.target.closest(".multi"))
    document.querySelectorAll(".pop").forEach(p => p.classList.remove("open"));
});
function pick(name, box) {
  box.checked ? SEL[name].add(box.value) : SEL[name].delete(box.value);
  labelMulti(name); render();
}
function toggleAll(name, on) { if (on) SEL[name].clear(); labelMulti(name);
  drawMulti(name, ""); render(); }
function labelMulti(name) {
  const n = SEL[name].size, cap = name[0].toUpperCase() + name.slice(1);
  document.getElementById("b_" + name).textContent =
    cap + ": " + (n === 0 ? "all" : (n === 1 ? [...SEL[name]][0].slice(0,22) : n + " selected"));
}
function clearFilters() {
  Object.keys(SEL).forEach(k => { SEL[k].clear(); labelMulti(k); drawMulti(k, ""); });
  document.getElementById("q").value = "";
  document.getElementById("deadline").checked = false;
  render();
}
function esc(s){ return (s||"").replace(/[&<>"]/g, c =>
  ({"&":"&amp;","<":"&lt;",">":"&gt;",'"':"&quot;"}[c])); }

function render() {
  const q = document.getElementById("q").value.toLowerCase();
  const type = document.getElementById("type").value;

  const dl = document.getElementById("deadline").checked;
  const fresh = document.getElementById("fresh").value;

  let rows = ROWS.filter(r =>
    (!type || r.Type === type) &&
    (SEL.priority.size === 0 || SEL.priority.has(r.Priority)) &&
    (SEL.country.size  === 0 || SEL.country.has(r.Country)) &&
    (SEL.company.size  === 0 || SEL.company.has(r.Company)) &&
    (!dl || r["Deadline found"]) &&
    (fresh === "" ||
     (fresh === "hideold" ? r.Freshness !== "Out of date" : r.Freshness === fresh)) &&
    (!q || (r.Company + " " + r["Role / posting"] + " " + (r.City||""))
             .toLowerCase().includes(q)));

  rows.sort((a,b) => ((a[sortKey]||"") > (b[sortKey]||"") ? 1 : -1) * sortDir);
  SHOWN = rows;

  document.querySelector("#tbl tbody").innerHTML = rows.map(r => `
    <tr>
      <td class="pick"><input type="checkbox" ${SAVED.has(rid(r))?"checked":""}
            onchange='toggleSave(${JSON.stringify(JSON.stringify(r))})'
            title="Add to my shortlist"></td>
      <td><span class="tag">${esc((r.Priority||"").replace(/ - /," "))}</span></td>
      <td><span class="co">${esc(r.Company)}</span>${r.City ? '<div class="muted">'+esc(r.City)+'</div>' : ''}</td>
      <td>${esc(r.Country)}</td>
      <td>${esc(r["Role / posting"])}</td>
      <td><span class="tag ${esc(r.Type)}">${esc(r.Type)}</span></td>
      <td>${esc(r["Deadline found"])}</td>
      <td>${esc((r.Posted||"").split(" (")[0])}
          <div class="muted fresh-${esc(r.Freshness).replace(/ /g,"")}">${esc(r.Freshness)}</div></td>
      <td class="muted">${esc(r["Date checked"])}</td>
      <td><a href="${esc(r.Link)}" target="_blank" rel="noopener">open</a>
          ${r["Social accounts"] ? " · " + r["Social accounts"].split(" | ")
             .map(u => `<a href="${esc(u)}" target="_blank" rel="noopener">social</a>`)
             .join(" ") : ""}</td>
    </tr>`).join("");
  document.getElementById("count").textContent =
    rows.length + " of " + ROWS.length + " rows";
}
function exportExcel() {
  const rows = SHOWN.map(r => ({...r}));
  fetch("/api/export/excel", {method:"POST",
      headers:{"Content-Type":"application/json"}, body: JSON.stringify(rows)})
    .then(r => r.blob()).then(b => {
      const a = document.createElement("a");
      a.href = URL.createObjectURL(b);
      a.download = "opera-jobs.xlsx";
      document.body.appendChild(a); a.click(); a.remove();
    });
}
function sortBy(k){ sortDir = (k === sortKey) ? -sortDir : 1; sortKey = k; render(); }

function startJob(key) {
  fetch("/api/run/" + key, {method:"POST"}).then(r => r.json()).then(d => {
    if (d.error) alert(d.error);
    poll();
  });
}
function poll() {
  fetch("/api/status").then(r => r.json()).then(s => {
    const log = document.getElementById("log");
    document.querySelectorAll("button").forEach(b => b.disabled = s.running);
    document.getElementById("status").textContent = s.running
      ? `Running: ${s.job} (started ${s.started}) — you can leave this page open`
      : (s.job ? `Last run: ${s.job}` : "Nothing running.");
    if (s.log.length) {
      log.style.display = "block";
      log.textContent = s.log.join("\\n");
      log.scrollTop = log.scrollHeight;
    }
    if (s.running) setTimeout(poll, 1500);
    else if (s.job) setTimeout(() =>
      fetch("/api/jobs").then(r => r.json()).then(d => { ROWS = d; render(); }), 500);
  });
}
poll();
</script>
"""


@app.route("/api/export/excel", methods=["POST"])
def api_export_excel():
    """Export exactly the rows the user is looking at, as a real .xlsx."""
    rows = request.get_json(force=True) or []
    cols = ["Priority", "Company", "City", "Country", "Role / posting", "Type",
            "Deadline found", "Posted", "Freshness", "Date checked", "Link",
            "Social accounts", "Found on page"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Findings"
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c, "") for c in cols])
    for i, w in enumerate([24, 28, 15, 9, 56, 12, 16, 13, 14, 13, 46, 40, 40],
                          start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="3B3833")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{max(ws.max_row, 2)}"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True, download_name="opera-jobs.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument."
                 "spreadsheetml.sheet")


@app.route("/api/saved")
def api_saved():
    return jsonify(tracker.load())


@app.route("/api/saved/add", methods=["POST"])
def api_saved_add():
    row = request.get_json(force=True) or {}
    added = tracker.add(row)
    return jsonify({"added": added, "id": tracker.row_id(row)})


@app.route("/api/saved/remove", methods=["POST"])
def api_saved_remove():
    d = request.get_json(force=True) or {}
    return jsonify({"removed": tracker.remove(d.get("id", ""))})


@app.route("/api/saved/update", methods=["POST"])
def api_saved_update():
    d = request.get_json(force=True) or {}
    ok = tracker.update(d.get("id", ""), d.get("field", ""), d.get("value", ""))
    return jsonify({"ok": ok})


@app.route("/api/saved/export")
def api_saved_export():
    path, _ = tracker.export_xlsx()
    return send_file(path, as_attachment=True,
                     download_name="my-opera-applications.xlsx")


@app.route("/")
def index():
    return render_template_string(PAGE, tasks=TASKS)


@app.route("/api/jobs")
def api_jobs():
    return jsonify(load_jobs())


@app.route("/api/status")
def api_status():
    with _lock:
        return jsonify(dict(state))


@app.route("/api/run/<key>", methods=["POST"])
def api_run(key):
    if key not in TASKS:
        return jsonify({"error": "unknown task"}), 400
    with _lock:
        if state["running"]:
            return jsonify({"error": "Something is already running."}), 409
    threading.Thread(target=run_task, args=(key,), daemon=True).start()
    return jsonify({"ok": True})


if __name__ == "__main__":
    print("\n  Opera Jobs app running.")
    print("  Open this in your browser:  http://localhost:5055")
    print("  Press Ctrl-C here to stop.\n")
    app.run(port=5055, debug=False)
