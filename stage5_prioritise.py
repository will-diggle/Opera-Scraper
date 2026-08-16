"""
STAGE 5 - sort the wheat from the chaff.

Operabase lists everyone: opera houses, but also amateur choral societies,
youth orchestras, church choirs, schools and agencies. They all use the word
"audition", so the raw results mix a chorus vacancy at Theater Kiel with a
village choir's open evening.

This adds a Priority column so you can filter to the ones that can actually
engage a professional singer. It does not delete anything - it just sorts.

Run:  .venv/bin/python stage5_prioritise.py
Out:  opera_jobs_final.xlsx
"""

import csv
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

JOBS = "jobs_targets.csv"
TARGETS = "targets.csv"
OUT = "opera_jobs_final.xlsx"


def rx(*words):
    return re.compile("|".join(words), re.I)


# Checked in order - first match wins.
CATEGORIES = [
    ("C - Youth / school", rx(
        "jugend", "kinder", "youth", "children", r"\bschul", "school",
        "universit", "student", "akademi", "academy", "conservat",
        "hochschule", "sängerknaben", "saengerknaben", "chorschule",
        "college", "junior", "scuola", "escuela", "école", "giovani")),
    ("C - Agency / coaching", rx(
        "agentur", "agency", "management", "artists? manage", "coaching",
        "künstleragentur", "kuenstleragentur", "casting agen")),
    ("C - Amateur / community choir", rx(
        "chorverein", "gesangverein", "choral society", "oratorienchor",
        "kirchenchor", "singkreis", "liederkranz", "männerchor", "maennerchor",
        "frauenchor", "chorgemeinschaft", "community choir", "amateur",
        "coro amatoriale", "cantorei", "kantorei", "singgemeinschaft",
        "choral union", "chorus society", "gospelchor", "shanty")),
    ("A - Opera house / theatre", rx(
        r"\boper", "opéra", "opera", "theater", "theatre", "teatro", "théâtre",
        "staatstheater", "landestheater", "stadttheater", "musiktheater",
        "lyric", "ópera")),
    ("A - Festival", rx(
        "festival", "festspiele", "festwochen", "biennale", "estate musicale")),
    ("B - Professional choir", rx(
        "chor", "chorus", "choir", "coro", "chœur", "choeur", "consort",
        "vocal ensemble", "cappella", "capella", "kammerchor", "rundfunkchor")),
    ("B - Orchestra / concert", rx(
        "orchester", "orchestra", "philharmon", "sinfoni", "symphony",
        "symphonie", "orquesta", "orchestre", "camerata", "ensemble")),
]

# Postings that are clearly not a singing engagement, whoever posted them.
NOT_FOR_YOU = rx(r"\bballet", r"\bballett", "danc", "ballerin", "violin",
                 "cello", "viola", "flute", "flöte", "oboe", "klarinett",
                 "clarinet", "fagott", "bassoon", "horn", "trumpet",
                 "trompete", "posaune", "trombone", "tuba", "percussion",
                 "schlagzeug", "harfe", "harp", "timpani", "pauke",
                 "korrepetit", "repetiteur", "dirigent", "conductor",
                 "orchestermusiker", "professor", "d'orchestra", "violon")


def categorise(name, org_type):
    blob = f"{name} {org_type}"
    for label, pattern in CATEGORIES:
        if pattern.search(blob):
            return label
    return "B - Other"


def main():
    with open(TARGETS, newline="", encoding="utf-8") as fh:
        meta = {r["name"]: r for r in csv.DictReader(fh)}
    with open(JOBS, newline="", encoding="utf-8") as fh:
        jobs = list(csv.DictReader(fh))

    for j in jobs:
        m = meta.get(j["Company"], {})
        j["Priority"] = categorise(j["Company"], m.get("org_type", ""))
        j["City"] = m.get("city", "")
        if NOT_FOR_YOU.search(j["Role / posting"]):
            j["Type"] = "Not singing"

    # A first, then B, then C; within a band, real singer postings first.
    jobs.sort(key=lambda r: (r["Priority"][0],
                             r["Type"] != "Singer",
                             r["Country"], r["Company"]))

    cols = ["Priority", "Company", "City", "Country", "Role / posting", "Type",
            "Deadline found", "Link", "Social accounts", "Found on page",
            "Date checked", "Status", "Notes"]
    widths = [26, 30, 16, 9, 58, 12, 16, 46, 40, 40, 12, 12, 26]

    wb = Workbook()
    ws = wb.active
    ws.title = "Opportunities"
    ws.append(cols)
    for j in jobs:
        ws.append([j.get(c, "") for c in cols])
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1]:
        c.font = Font(bold=True)
    green = PatternFill("solid", fgColor="C6EFCE")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[0].value.startswith("A") and row[5].value == "Singer":
            row[0].fill = green
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
    wb.save(OUT)

    from collections import Counter
    print(f"Wrote {OUT}: {len(jobs)} rows")
    for k, v in sorted(Counter(j["Priority"] for j in jobs).items()):
        singers = sum(1 for j in jobs
                      if j["Priority"] == k and j["Type"] == "Singer")
        print(f"  {k:32} {v:5} rows  ({singers} singer)")


if __name__ == "__main__":
    main()
